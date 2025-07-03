import time
from multiprocessing import Process
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from webdriver_manager.firefox import GeckoDriverManager
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.edge.service import Service as EdgeService

wb = load_workbook(r"C:\Users\Arcknight\Desktop\HCL_Testing_Training\Coding\Excel\cities.xlsx")
sheet = wb.active
rows = list(sheet.iter_rows(min_row=2, max_col=2, values_only=True))

def run_browser(browser_name, rows_to_process):
    if browser_name == "firefox":
        service = FirefoxService(GeckoDriverManager().install())
        driver = webdriver.Firefox(service=service)
    elif browser_name == "edge":
        service = EdgeService(EdgeChromiumDriverManager().install())
        driver = webdriver.Edge(service=service)
    else:
        raise ValueError("Unsupported browser!")

    wait = WebDriverWait(driver, 20)
    driver.maximize_window()

    for row in rows_to_process:
        from_city, to_city = row
        driver.get("https://www.cleartrip.com/")
        time.sleep(2)

        from_input = wait.until(EC.element_to_be_clickable((By.XPATH,
            "/html/body/div[1]/div/main/div/div/div/div[1]/div[1]/div/div[1]/div[2]/div/div[2]/div/div[1]/input"
        )))
        from_input.clear()
        from_input.send_keys(from_city)
        time.sleep(1)
        wait.until(EC.element_to_be_clickable((By.XPATH,
            "/html/body/div[1]/div/main/div/div/div/div[1]/div[1]/div/div[1]/div[2]/div/div[2]/div[1]/div[1]/div[2]/ul/li[1]"
        ))).click()

        to_input = wait.until(EC.element_to_be_clickable((By.XPATH,
            "/html/body/div[1]/div/main/div/div/div/div[1]/div[1]/div/div[1]/div[2]/div/div[2]/div/div[3]/input"
        )))
        to_input.clear()
        to_input.send_keys(to_city)
        time.sleep(1)
        wait.until(EC.element_to_be_clickable((By.XPATH,
            "/html/body/div[1]/div/main/div/div/div/div[1]/div[1]/div/div[1]/div[2]/div/div[2]/div[1]/div[3]/div[2]/ul/li[1]"
        ))).click()

        dep_open = wait.until(EC.element_to_be_clickable((By.XPATH,
            "/html/body/div[1]/div/main/div/div/div/div[1]/div[1]/div/div[1]/div[2]/div/div[4]/div/div/div/div[1]/div[1]"
        )))
        dep_open.click()
        time.sleep(1)
        wait.until(EC.element_to_be_clickable((By.XPATH,
            "//div[@aria-label='Mon Jun 30 2025']"
        ))).click()
        time.sleep(1)

        return_open = wait.until(EC.element_to_be_clickable((By.XPATH,
            "/html/body/div[1]/div/main/div/div/div/div[1]/div[1]/div/div[1]/div[2]/div/div[4]/div/div/div/div[3]"
        )))
        return_open.click()
        time.sleep(1)

        wait.until(EC.element_to_be_clickable((By.XPATH,
            "/html/body/div[1]/div/main/div/div/div/div[1]/div[1]/div/div[1]/div[2]/div/div[4]/div/div/div/div[4]/div[3]/div/div[2]/div[2]/div[3]/div[1]/div[6]/div/div"
        ))).click()
        time.sleep(1)

        wait.until(EC.element_to_be_clickable((By.XPATH,
            "//button[.='Search flights']"
        ))).click()

        time.sleep(8)

    driver.quit()

# Start both browsers - each does full task
if __name__ == "__main__":
    p1 = Process(target=run_browser, args=("firefox", rows))
    p2 = Process(target=run_browser, args=("edge", rows))

    p1.start()
    time.sleep(2)
    p2.start()

    p1.join()
    p2.join()
