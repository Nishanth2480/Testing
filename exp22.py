import time
from multiprocessing import Process
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.firefox import GeckoDriverManager
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.edge.service import Service as EdgeService
from openpyxl import load_workbook

wb = load_workbook(r"C:\Users\Arcknight\Desktop\HCL_Testing_Training\Coding\Excel\cities.xlsx")
sheet = wb.active
rows = list(sheet.iter_rows(min_row=2, max_col=2, values_only=True))

def run_browser(browser_name, rows_to_process):
    if browser_name == "firefox":
        service = FirefoxService(GeckoDriverManager().install())
        driver = webdriver.Firefox(service=service)
    elif browser_name == "edge":
        service = EdgeService(EdgeChromiumDriverManager().install())
        driver = webdriver.Edge(service=service)
    else:
        raise ValueError("Unsupported browser!")

    wait = WebDriverWait(driver, 20)
    driver.maximize_window()

    for row in rows_to_process:
        from_city, to_city = row
        driver.get("https://www.goibibo.com/flights/")
        time.sleep(5)

        # Close any pop-ups
        try:
            driver.find_element(By.CSS_SELECTOR, "[data-testid='loginPopup-closeBtn']").click()
        except:
            pass

        # Click and enter From city
        from_input = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/div[4]/div/div/div[2]/div[1]/div/div[2]/div/input")))
        from_input.click()
        from_input.clear()
        from_input.send_keys(from_city)
        time.sleep(2)
        wait.until(EC.element_to_be_clickable((By.XPATH, "//ul/li[1]/div"))).click()

        # Click and enter To city
        to_input = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/div[4]/div/div/div[2]/div[2]/div/div[2]/div/input")))
        to_input.click()
        to_input.clear()
        to_input.send_keys(to_city)
        time.sleep(2)
        wait.until(EC.element_to_be_clickable((By.XPATH, "//ul/li[1]/div"))).click()

        # Select departure date
        dep_date = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/div[4]/div/div/div[2]/div[3]/div")))
        dep_date.click()
        time.sleep(1)
        wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@aria-label='Mon Jun 30 2025']"))).click()

        # Select return date
        return_date = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/div[4]/div/div/div[2]/div[4]/div")))
        return_date.click()
        time.sleep(1)
        wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@aria-label='Wed Jul 02 2025']"))).click()

        # Click Search
        wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/div[4]/div/div/div[4]/span"))).click()

        time.sleep(8)

    driver.quit()

if __name__ == "__main__":
    p1 = Process(target=run_browser, args=("firefox", rows))
    p2 = Process(target=run_browser, args=("edge", rows))

    p1.start()
    time.sleep(2)
    p2.start()

    p1.join()
    p2.join()
