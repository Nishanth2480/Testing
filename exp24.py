import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.edge.service import Service as EdgeService
from openpyxl import load_workbook

browser = "firefox"

# Set up driver based on chosen browser
if browser == "chrome":
    service = ChromeService("C:\\WebDriver\\chromedriver-win64\\chromedriver.exe")
    driver = webdriver.Chrome(service=service)
elif browser == "firefox":
    service = FirefoxService("C:\\WebDriver\\geckodriver-win64\\geckodriver.exe")
    driver = webdriver.Firefox(service=service)
elif browser == "edge":
    service = EdgeService("C:\\WebDriver\\edgedriver-win64\\msedgedriver.exe")
    driver = webdriver.Edge(service=service)
else:
    raise Exception("Unsupported browser")

driver.maximize_window()

# Load Excel data
wb = load_workbook(r"C:\Users\Arcknight\Desktop\HCL_Testing_Training\Coding\Excel\cities.xlsx")
sheet = wb.active

for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
    from_city, to_city = row

    driver.get("https://www.makemytrip.com/")
    time.sleep(5)

    # Dismiss login popup if present
    try:
        driver.find_element(By.XPATH, "//span[@class='commonModal__close']").click()
    except:
        pass

    # FROM CITY
    driver.find_element(By.XPATH, "/html/body/div/div/div[2]/div/div/div/div/div/div[2]/div[1]/div[1]/label").click()
    time.sleep(1)
    from_input = driver.find_element(By.XPATH, "/html/body/div/div/div[2]/div/div/div/div/div/div[2]/div[1]/div[1]/div[1]/div/div/div/input")
    from_input.send_keys(from_city)
    time.sleep(2)
    from_input.send_keys(Keys.DOWN,Keys.ENTER)

    # TO CITY
    to_input = driver.find_element(By.XPATH, "/html/body/div/div/div[2]/div/div/div/div/div/div[2]/div[1]/div[2]/div[1]/div/div/div[1]/input")
    to_input.send_keys(to_city)
    time.sleep(2)
    to_input.send_keys(Keys.DOWN,Keys.ENTER)

    # DEPARTURE DATE
    driver.find_element(By.XPATH, "/html/body/div/div/div[2]/div/div/div/div/div/div[2]/div[1]/div[3]").click()
    time.sleep(1)
    driver.find_element(By.XPATH, "/html/body/div/div/div[2]/div/div/div/div/div/div[2]/div[1]/div[3]/div[1]/div/div/div/div[2]/div/div[2]/div[1]/div[3]/div[5]/div[2]/div").click()

    # RETURN DATE
    driver.find_element(By.XPATH, "/html/body/div/div/div[2]/div/div/div/div/div/div[2]/div[1]/div[4]/div[2]").click()
    time.sleep(1)
    driver.find_element(By.XPATH, "/html/body/div/div/div[2]/div/div/div/div/div/div[2]/div[1]/div[3]/div[1]/div/div/div/div[2]/div/div[2]/div[2]/div[3]/div[1]/div[7]").click()

    # Click Search
    time.sleep(1)
    driver.find_element(By.XPATH, "/html/body/div/div/div[2]/div/div/div/div/div/div[2]/p/a").click()

    # Wait to load results
    time.sleep(10)

driver.quit()
