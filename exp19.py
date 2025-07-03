import time
import openpyxl as XL
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.devtools.v135.browser import close
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait

path=r"C:\Users\Arcknight\Desktop\HCL_Testing_Training\Coding\Excel\Book1.xlsx"
w_b=XL.load_workbook(path)
sheet=w_b.active
service = Service(r"C:\WebDriver\chromedriver-win64\chromedriver.exe")
driver = webdriver.Chrome(service=service)
driver.maximize_window()
wait = WebDriverWait(driver, 10)
driver.get("https://www.cleartrip.com/")
time.sleep(3)

try:
    close_pop_up=driver.find_element(By.XPATH,'/html/body/div[2]/div/div/div/div/div[2]/div/div[1]/div[2]')
    close_pop_up.click()
    print('popUpClosed')
    time.sleep(2)
except:
    print('Not found')


row=sheet.max_row
col=sheet.max_column

for i in range(2,row+1):
    from_value,to_value=sheet.cell(row=i,column=1).value,sheet.cell(row=i,column=2).value
    from_field = driver.find_element(By.XPATH, '//input[@placeholder="Where from?"]')
    from_field.click()
    time.sleep(1)
    from_check,to_check=False,False
    options = driver.find_elements(By.CSS_SELECTOR, "ul.airportList > li")
    for option in options:
        curr_value=option.find_element(By.CSS_SELECTOR,"p").text.strip()
        print(curr_value,from_value,curr_value==from_value)
        if curr_value == from_value:
            option.click()
            time.sleep(1)
            print("Successfully From filled")
            from_check=True
            break

    to_field = driver.find_element(By.XPATH, '//input[@placeholder="Where to?"]')
    to_field.click()
    time.sleep(1)
    options = driver.find_elements(By.CSS_SELECTOR, "ul.airportList > li")
    for option in options:
        curr_value = option.find_element(By.CSS_SELECTOR, "p").text.strip()
        if curr_value == to_value:
            option.click()
            time.sleep(1)
            print(curr_value, from_value, curr_value == from_value)

            print("Successfully To filled")
            to_check=True
            break
    if from_check and to_check:
        return_date = wait.until(EC.element_to_be_clickable((
            By.CSS_SELECTOR, 'div[data-testid="dateSelectReturn"]'
        )))
        return_date.click()
        time.sleep(1)
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", return_date)
        time.sleep(1)
        driver.find_element(By.XPATH, '//div[@class="flex-1 ta-right"]//*[name()="svg"]').click()
        time.sleep(2)

        driver.find_element(By.XPATH,
                            "//div[contains(@aria-label,'Sat Aug 30 2025')]//div[contains(@class,'day-gridContent') and not(contains(@class,'selected'))]").click()

        time.sleep(1)

        add_student_fare = driver.find_element(By.XPATH, '//p[normalize-space()="Student fare"]')
        add_student_fare.click()

        time.sleep(1)

        add_student_fare.click()

        time.sleep(1)

        add_student_fare.click()

        search_flights = driver.find_element(By.XPATH, '//h4[normalize-space()="Search flights"]')
        search_flights.click()
        time.sleep(3)


        try:
            book_wait=WebDriverWait(driver,8)
            book = book_wait.until(
                EC.presence_of_element_located((By.XPATH, '//p[@class="sc-eqUAAy gqwtSu c-pointer fs-14 lh-20"]')))
            # book=driver.find_element(By.XPATH,'//p[@class="sc-eqUAAy gqwtSu c-pointer fs-14 lh-20"]')
            book.click()
        except:
            print("No flights are avilable")

        time.sleep(1)

        driver.back()
        time.sleep(0.5)
        driver.refresh()
        time.sleep(1)
    else:
        driver.refresh()
        continue

title=driver.title
print(title)
print(driver.current_url)
driver.quit()