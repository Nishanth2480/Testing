import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.edge.service import Service as EdgeService
from openpyxl import load_workbook

browser = "firefox"

if browser == "chrome":
    service = ChromeService("C:\\WebDriver\\chromedriver-win64\\chromedriver.exe")
    driver = webdriver.Chrome(service=service)
elif browser == "firefox":
    service = FirefoxService("C:\\WebDriver\\geckodriver-win64\\geckodriver.exe")
    driver = webdriver.Firefox(service=service)
elif browser == "edge":
    service = EdgeService("C:\\WebDriver\\edgedriver-win64\\msedgedriver.exe")
    driver = webdriver.Edge(service=service)
else:
    raise Exception("Unsupported browser")

driver.maximize_window()

wb = load_workbook(r"C:\Users\Arcknight\Desktop\HCL_Testing_Training\Coding\Excel\cities.xlsx")
sheet = wb.active

for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
    from_city, to_city = row

    # Refresh page state completely
    driver.get("https://www.goibibo.com/flights/")
    driver.delete_all_cookies()
    time.sleep(5)

    # FROM
    driver.find_element(By.XPATH, "/html/body/div/div[4]/div/div/div[2]/div[1]/div/div/p").click()
    time.sleep(1)
    from_input = driver.find_element(By.XPATH, "/html/body/div/div[4]/div/div/div[2]/div[1]/div/div[2]/div/input")
    from_input.send_keys(from_city)
    time.sleep(2)
    from_input.send_keys(Keys.ENTER)

    # TO
    to_input = driver.find_element(By.XPATH, "/html/body/div/div[4]/div/div/div[2]/div[2]/div/div[2]/div/input")
    to_input.send_keys(to_city)
    time.sleep(2)
    to_input.send_keys(Keys.DOWN,Keys.ENTER)

    # Departure date
    driver.find_element(By.XPATH, "/html/body/div/div[4]/div/div/div[2]/div[3]/div").click()
    time.sleep(1)
    driver.find_element(By.XPATH,
        "/html/body/div/div[4]/div/div/div[2]/div[3]/div[2]/div[2]/div/div/div[2]/div[1]/div[3]/div[5]/div[2]").click()

    # Return date
    time.sleep(1)
    driver.find_element(By.XPATH, "/html/body/div/div[4]/div/div/div[2]/div[4]/div").click()
    time.sleep(1)
    driver.find_element(By.XPATH,
        "/html/body/div/div[4]/div/div/div[2]/div[4]/div[2]/div[2]/div/div/div[2]/div[1]/div[3]/div[1]/div[7]").click()

    # ⛏️ Close the return calendar popup
    time.sleep(1)
    try:
        driver.find_element(By.XPATH,
            "/html/body/div/div[4]/div/div/div[2]/div[4]/div[2]/div[1]/span").click()
    except:
        pass

    # Click Search
    time.sleep(2)
    driver.find_element(By.XPATH, "/html/body/div/div[4]/div/div/div[4]/span").click()

    time.sleep(10)

driver.quit()
