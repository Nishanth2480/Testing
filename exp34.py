import unittest
import time
import psutil
import logging
import os
import HtmlTestRunner
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

class OrangeHRMLoginTest(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.report_folder = r"C:\Users\Arcknight\Desktop\HCL_Testing_Training\Coding"
        cls.log_path = os.path.join(cls.report_folder, "log", "OrangeHRM_login.log")
        os.makedirs(os.path.dirname(cls.log_path), exist_ok=True)
        logging.basicConfig(filename=cls.log_path, level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
        logging.info("=== Starting OrangeHRM Test Suite ===")

        cls.excel_path = os.path.join(cls.report_folder, "reports", "login_test_results.xlsx")
        os.makedirs(os.path.dirname(cls.excel_path), exist_ok=True)
        if not os.path.exists(cls.excel_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"
            ws.append(["Test Name", "Status", "Message"])
            wb.save(cls.excel_path)

    def log_to_excel(self, test_name, status, message):
        wb = load_workbook(self.excel_path)
        ws = wb["Results"]
        ws.append([test_name, status, message])
        wb.save(self.excel_path)

    def log_to_txt(self, test_name, status, message):
        msg = f"{test_name} - {status}: {message}"
        logging.info(msg) if status == "PASS" else logging.error(msg)

    def log_result(self, test_name, status, message=""):
        self.log_to_excel(test_name, status, message)
        self.log_to_txt(test_name, status, message)

    def setUp(self):
        self.service = Service(r"C:\WebDriver\chromedriver-win64\chromedriver.exe")
        self.driver = webdriver.Chrome(service=self.service)
        self.driver.maximize_window()
        self.driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
        self.wait = WebDriverWait(self.driver, 10)

    def show_alert(self, message):
        self.driver.execute_script(f"alert('{message}');")
        time.sleep(2)
        self.driver.switch_to.alert.accept()

    def test_1_valid_login(self):
        test_name = "Valid Login"
        try:
            self.wait.until(EC.presence_of_element_located((By.NAME, "username"))).send_keys("Admin")
            time.sleep(2)
            self.wait.until(EC.presence_of_element_located((By.NAME, "password"))).send_keys("admin123")
            time.sleep(2)
            self.wait.until(EC.element_to_be_clickable((By.XPATH,"/html/body/div/div[1]/div/div[1]/div/div[2]/div[2]/form/div[3]/button"))).click()
            time.sleep(2)
            self.wait.until(EC.presence_of_element_located((By.XPATH, "//h6[text()='Dashboard']")))

            assert "dashboard" in self.driver.current_url.lower(), "Dashboard not in URL"
            assert "orangehrm" in self.driver.title.lower(), "Title doesn't contain orangehrm"
            assert self.driver.find_element(By.XPATH, "//h6").text.strip().lower() == "dashboard", "Dashboard heading not found"

            self.show_alert("True")
            self.log_result(test_name, "PASS", "Login successful and Dashboard loaded.")
        except AssertionError as e:
            self.show_alert("False")
            self.log_result(test_name, "FAIL", str(e))
            raise e

    def test_2_invalid_login(self):
        test_name = "Invalid Login"
        try:
            self.wait.until(EC.presence_of_element_located((By.NAME, "username"))).send_keys("Admin")
            time.sleep(2)
            self.wait.until(EC.presence_of_element_located((By.NAME, "password"))).send_keys("Dark$2004")
            time.sleep(2)
            self.wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/div[1]/div/div[1]/div/div[2]/div[2]/form/div[3]/button"))).click()
            time.sleep(2)

            error_element = self.wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Invalid credentials')]")))
            assert "invalid" in error_element.text.lower(), "Invalid credentials error not shown"

            self.show_alert("True")
            self.log_result(test_name, "PASS", "Proper error message shown.")
        except Exception as e:
            screenshot_dir = r"C:\Users\Arcknight\Desktop\HCL_Testing_Training\Coding\Screenshot"
            os.makedirs(screenshot_dir, exist_ok=True)
            self.driver.save_screenshot(f"{screenshot_dir}\\{test_name.replace(' ', '_')}_error.png")
            self.show_alert("False")
            self.log_result(test_name, "FAIL", str(e))
            raise e


    def test_3_empty_login(self):
        test_name = "Empty Login"
        try:
            self.wait.until(EC.presence_of_element_located((By.NAME, "username"))).send_keys("")
            time.sleep(2)
            self.wait.until(EC.presence_of_element_located((By.NAME, "password"))).send_keys("")
            time.sleep(2)
            self.wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/div[1]/div/div[1]/div/div[2]/div[2]/form/div[3]/button"))).click()
            time.sleep(2)

            required_elements = self.wait.until(EC.presence_of_all_elements_located((By.XPATH, "//*[contains(text(),'Required')]")))
            assert any("required" in el.text.lower() for el in required_elements), "Required message not shown"

            self.show_alert("True")
            self.log_result(test_name, "PASS", "Required field message displayed.")
        except Exception as e:
            screenshot_dir = r"C:\Users\Arcknight\Desktop\HCL_Testing_Training\Coding\Screenshot"
            os.makedirs(screenshot_dir, exist_ok=True)
            self.driver.save_screenshot(f"{screenshot_dir}\\{test_name.replace(' ', '_')}_error.png")
            self.show_alert("False")
            self.log_result(test_name, "FAIL", str(e))
            raise e


    def tearDown(self):
        self.driver.quit()
        for proc in psutil.process_iter(['pid', 'name']):
            if 'chromedriver' in proc.info['name'].lower():
                proc.kill()

    @classmethod
    def tearDownClass(cls):
        logging.info("=== OrangeHRM Test Suite Completed ===\n")


if __name__ == "__main__":
    unittest.main(
        testRunner=HtmlTestRunner.HTMLTestRunner(
            output=os.path.join(r"C:\Users\Arcknight\Desktop\HCL_Testing_Training\Coding", "reports"),
            report_name="orangehrm_login_report",
            report_title="OrangeHRM Login Test Report"
        )
    )